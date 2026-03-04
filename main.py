import streamlit as st
import os
import time
from openpyxl import Workbook, load_workbook

# Konfiguration der Seite
st.set_page_config(page_title="Registrierung", page_icon="📝")

excel_workbook = "core_database.xlsx"


# --- LOGIK: SPEICHERN IN EXCEL ---
def save_to_excel(name, email, passwort, geschlecht, kommentar, alter):
    if os.path.exists(excel_workbook):
        wb = load_workbook(excel_workbook)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["User ID", "Email", "Key", "Origin", "Log", "Alter"])

    ws.append([name, email, passwort, geschlecht, kommentar, alter])
    wb.save(excel_workbook)


# --- UI: FORMULAR ---
st.title("📝 Registrierungsformular")

# Ein Container für das Formular (ähnlich wie dein HTML-Template)
with st.form("registration_form"):
    username = st.text_input("Username")
    email = st.text_input("Email")
    passwort = st.text_input("Passwort", type="password")
    geschlecht = st.selectbox("Geschlecht", ["Männlich", "Weiblich", "Divers"])
    alter = st.number_input("Alter", min_value=0, max_value=120, value=18)
    kommentar = st.text_area("Kommentar")

    # Der Submit-Button
    submit_button = st.form_submit_button("Registrieren")

if submit_button:
    if username and email:  # Einfache Validierung
        with st.spinner("Wird gespeichert..."):
            time.sleep(0.8)  # Dein Delay
            save_to_excel(username, email, passwort, geschlecht, kommentar, alter)

        st.success(f"Erfolg! Hallo {username}, deine Daten wurden gespeichert.")
        st.balloons()
    else:
        st.error("Bitte fülle mindestens Name und Email aus.")
