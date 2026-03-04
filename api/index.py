from flask import Flask, request, render_template, url_for
import os, time
from openpyxl import Workbook, load_workbook


app = Flask(__name__)
excel_workbook = "core_database.xlsx"

@app.route("/")
def index():
    return render_template("registrierungsformular.html")

@app.route("/submit", methods=["POST"])
def submit():
    time.sleep(0.8)
    name = request.form.get("username")
    email = request.form.get("email")
    passwort = request.form.get("passwort")
    geschlecht = request.form.get("geschlecht")
    kommentar = request.form.get("kommentar")
    alter = request.form.get("alter")
    theme = request.form.get("active_theme_input") or "light"

    if os.path.exists(excel_workbook):
        wb = load_workbook(excel_workbook)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["User ID", "Email", "Key", "Origin", "Log", "Alter"])

    ws.append([name, email, passwort, geschlecht, kommentar, alter])
    wb.save(excel_workbook)
    return render_template("success.html", name=name, theme=theme)

if __name__ == "__main__":
    app.run(debug=True)
